import os
import json
import logging
import random
import string
from fastapi import FastAPI, Request, Header, HTTPException
from pydantic import BaseModel
from telegram import (
    Bot,
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    WebAppInfo,
)
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    CallbackQueryHandler,
    filters,
)
import asyncio
from typing import Optional, Dict, List, Any
from enum import Enum
import requests

# Base URL for the restaurant API.  The bot will query this endpoint to retrieve
# the list of restaurants instead of reading a local file.  This can be
# overridden via the SERVER_API_BASE_URL environment variable.  By default it
# assumes the Node.js server is running on localhost:3000.
SERVER_API_BASE_URL = os.getenv('SERVER_API_BASE_URL', 'http://localhost:3000')

# Settings
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '8157721863:AAF2WwoxcOlDn6L-Tl_3-VSGer5QTPvkwYM')
FASTAPI_SECRET_TOKEN = os.getenv('FASTAPI_SECRET_TOKEN', 'supersecret')
PORT = int(os.getenv('PORT', 8000))
# Load super admin from environment and additional admins from file
# The first ID in the ADMIN_IDS environment variable is considered the super admin.
# Any additional admin IDs are persisted in admins.json.  The combined set defines all admins.
ADMIN_IDS_ENV = json.loads(os.getenv('ADMIN_IDS', '[6125606244]'))
if not isinstance(ADMIN_IDS_ENV, list):
    ADMIN_IDS_ENV = [ADMIN_IDS_ENV]

SUPER_ADMIN_ID = ADMIN_IDS_ENV[0]  # The super admin defined in code

# Path to store additional admin IDs (besides the super admin)
ADMINS_FILE = "admins.json"

def load_admin_list():
    try:
        if os.path.exists(ADMINS_FILE):
            with open(ADMINS_FILE, "r") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
    except Exception as e:
        logger.error(f"Error loading {ADMINS_FILE}: {e}")
    return []

def save_admin_list(admin_list):
    try:
        with open(ADMINS_FILE, "w") as f:
            json.dump(admin_list, f)
    except Exception as e:
        logger.error(f"Error saving {ADMINS_FILE}: {e}")

EXTRA_ADMIN_IDS = set(load_admin_list())

USERS_FILE = "subscribed_users.json"
COURIERS_FILE = "couriersr.json"
KITCHENS_FILE = "kitchensr.json"
MESSAGES_FILE = "messagesr.json"  # Legacy messages file (no longer used)

# Completed deliveries storage file.  Each entry contains keys: order_id, courier_id, delivered_at (ISO
# string with timezone), and delivery_fee.  This file is used to generate daily courier statistics.
DELIVERIES_FILE = "deliveries.json"

# Logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("bot")

# Data models
class UserRole(str, Enum):
    CUSTOMER = "customer"
    COURIER = "courier"
    KITCHEN = "kitchen"
    ADMIN = "admin"

class UserStatus(str, Enum):
    FREE = "free"
    BUSY = "busy"
    UNAVAILABLE = "unavailable"

class ContainerModel(BaseModel):
    name: str
    quantity: int
    price_per_unit: float
    total_price: float

class CustomerModel(BaseModel):
    name: str
    phone: str
    address: str
    note: str = ""
    location: Optional[Dict[str, float]] = None

class OrderItemModel(BaseModel):
    name: str
    quantity: int
    price_per_unit: float
    total_price: float
    restaurant: str
    container: Optional[ContainerModel] = None

class TotalsModel(BaseModel):
    dishes: float
    containers: float
    delivery: float
    discount: float
    final: float

class OrderModel(BaseModel):
    order_id: int
    status: str
    created_at: str
    customer: CustomerModel
    payment_method: str
    items: List[OrderItemModel]
    totals: TotalsModel

class CourierModel(BaseModel):
    id: int
    name: str
    # Optional phone number for the courier.  Couriers must provide this before accepting orders.
    phone: Optional[str] = None
    status: UserStatus = UserStatus.FREE
    current_order: Optional[int] = None

class KitchenModel(BaseModel):
    name: str
    admin_id: int

class MessageModel(BaseModel):
    from_user: int
    to_admin: bool
    text: str
    timestamp: str

# Data storage functions
def load_data(filename, default):
    try:
        if os.path.exists(filename):
            with open(filename, "r") as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Error loading {filename}: {e}")
    return default()

def save_data(filename, data):
    try:
        with open(filename, "w") as f:
            json.dump(data, f)
    except Exception as e:
        logger.error(f"Error saving {filename}: {e}")

# -----------------------------------------------------------------------------
# User subscription management
#
# Previously the bot only stored a list of chat IDs in subscribed_users.json and
# transiently held phone numbers in a runtime dictionary.  This meant that
# whenever the bot restarted it would forget users' phone numbers and prompt
# them again.  To persist user metadata (id, phone, username, full name) we
# store a list of dictionaries in subscribed_users.json.  Each entry has at
# minimum the `id` field, and may also include `phone`, `name` and `username`.
#
# Helper functions to load and save this structure.  When loading, if the file
# contains only a list of integers (legacy format), it will be converted into
# the new format with unknown phone/name/username fields.

def load_subscribed_users() -> List[Dict[str, Any]]:
    data = load_data(USERS_FILE, list)
    users: List[Dict[str, Any]] = []
    # If the JSON is already a list of dicts with id field, use as-is
    if data and isinstance(data, list) and all(isinstance(item, dict) for item in data):
        for item in data:
            # Ensure id is present and int
            if 'id' in item:
                try:
                    item['id'] = int(item['id'])
                except Exception:
                    continue
                users.append(item)
    else:
        # Legacy format: list of ints (chat IDs).  Convert to new structure.
        for entry in data:
            try:
                uid = int(entry)
                users.append({'id': uid})
            except Exception:
                continue
    return users


def save_subscribed_users(users: List[Dict[str, Any]]):
    # Persist list of user dictionaries into subscribed_users.json
    try:
        with open(USERS_FILE, 'w') as f:
            json.dump(users, f)
    except Exception as e:
        logger.error(f"Error saving {USERS_FILE}: {e}")


# Load initial data
# Load subscribed users as a list of dicts.  user_chats will derive from these
subscribed_users: List[Dict[str, Any]] = load_subscribed_users()
# Create a set of chat IDs for quick broadcast operations
user_chats: set[int] = set(u['id'] for u in subscribed_users if 'id' in u)
couriers = [CourierModel(**c) for c in load_data(COURIERS_FILE, list)]
kitchens = [KitchenModel(**k) for k in load_data(KITCHENS_FILE, list)]
# messages list is no longer used for persistence; we initialize it empty.  If
# there are entries in messagesr.json from previous runs they are ignored.
messages: List[MessageModel] = []
# Build phone mapping from subscribed_users.  Each phone maps to a user id.
phone_mapping: Dict[str, int] = {}
for u in subscribed_users:
    phone = u.get('phone')
    uid = u.get('id')
    if phone and uid is not None:
        phone_mapping[phone] = uid

# Load completed deliveries.  Each item is a dict as described above.
deliveries: List[Dict[str, Any]] = load_data(DELIVERIES_FILE, list)

# Active orders mapping: order_id -> {
#   'customer_id': user_id or None,
#   'courier_id': courier_id or None,
#   'location': Optional[Dict[str, float]]
# }
# This dictionary is used to track which courier and customer are associated with each order
active_orders: Dict[int, Dict[str, Optional[int]]] = {}

# Telegram bot setup
bot = Bot(TELEGRAM_BOT_TOKEN)
application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

# Helper functions
def is_admin(user_id: int) -> bool:
    """
    Determine whether a user is an admin.  The super admin is defined by SUPER_ADMIN_ID,
    while additional admins are stored in EXTRA_ADMIN_IDS.  This function returns True
    if the user is either the super admin or one of the extra admins.
    """
    return user_id == SUPER_ADMIN_ID or user_id in EXTRA_ADMIN_IDS

def get_courier(user_id: int) -> Optional[CourierModel]:
    for courier in couriers:
        if courier.id == user_id:
            return courier
    return None

def get_kitchen(admin_id: int) -> Optional[KitchenModel]:
    for kitchen in kitchens:
        if kitchen.admin_id == admin_id:
            return kitchen
    return None

def get_available_couriers() -> List[CourierModel]:
    # Only return couriers that are free and have provided a phone number
    return [c for c in couriers if c.status == UserStatus.FREE and c.phone]

def assign_order_to_courier(order_id: int) -> Optional[CourierModel]:
    available = get_available_couriers()
    if not available:
        return None
    
    courier = available[0]
    courier.status = UserStatus.BUSY
    courier.current_order = order_id
    save_data(COURIERS_FILE, [c.dict() for c in couriers])
    return courier

def get_courier_by_order(order_id: int) -> Optional[CourierModel]:
    for courier in couriers:
        if courier.current_order == order_id:
            return courier
    return None

def find_user_by_phone(phone: str) -> Optional[int]:
    """
    Given a phone string, attempt to find a matching user_id.
    The comparison ignores any non‑digit characters and matches based on the last nine digits.
    This allows both 994286407 and +998994286407 to map to the same user.
    """
    if not phone:
        return None
    # Extract digits only
    digits = ''.join(filter(str.isdigit, phone))
    # Use last nine digits for matching
    if len(digits) >= 9:
        digits = digits[-9:]
    for stored_phone, user_id in phone_mapping.items():
        stored_digits = ''.join(filter(str.isdigit, stored_phone))
        if len(stored_digits) >= 9:
            stored_digits = stored_digits[-9:]
        if stored_digits == digits:
            return user_id
    return None

def normalize_phone_input(input_value: str) -> Optional[str]:
    """
    Normalize a phone input by stripping non‑digit characters and removing the country code if present.
    Accepts a phone in two forms: 9 digits (e.g. 991234567) or +998 followed by 9 digits (e.g. +998991234567).
    Returns the normalized 9‑digit string if valid, otherwise None.
    """
    if not input_value:
        return None
    digits = ''.join(filter(str.isdigit, input_value))
    # Remove leading 998 if present (for +998 country code)
    if digits.startswith('998'):
        digits = digits[3:]
    # Accept exactly 9 digits
    if len(digits) == 9:
        return digits
    return None

async def send_to_admins(message: str, **kwargs):
    """
    Send a text message to all admin users (super and extra admins).  Accepts any additional
    keyword arguments for the Telegram send_message API.
    """
    for admin_id in [SUPER_ADMIN_ID] + list(EXTRA_ADMIN_IDS):
        try:
            await bot.send_message(chat_id=admin_id, text=message, **kwargs)
        except Exception as e:
            logger.error(f"Failed to send to admin {admin_id}: {e}")

async def send_to_couriers(message: str, exclude: Optional[int] = None, **kwargs):
    for courier in couriers:
        if courier.id != exclude and courier.status == UserStatus.FREE:
            try:
                await bot.send_message(chat_id=courier.id, text=message, **kwargs)
            except Exception as e:
                logger.error(f"Failed to send to courier {courier.id}: {e}")

async def send_to_kitchen(restaurant: str, message: str, **kwargs):
    # Match the restaurant name against the kitchen name case-insensitively and ignoring leading/trailing spaces.
    for kitchen in kitchens:
        if kitchen.name and restaurant and kitchen.name.strip().lower() == restaurant.strip().lower():
            try:
                await bot.send_message(chat_id=kitchen.admin_id, text=message, **kwargs)
            except Exception as e:
                logger.error(f"Failed to send to kitchen {kitchen.name}: {e}")
            break

# Order re‑assignment utilities
async def check_pending_orders_for_courier(courier: CourierModel):
    """
    Inspect all active orders that do not yet have a courier assigned (courier_id is None) and
    send the order details to the specified courier if they are available (status FREE and
    have provided a phone number).  After sending an order to this courier, schedule an
    automatic assignment task in 15 seconds so that if the courier does not accept the
    order in that time frame, it may be randomly assigned.
    """
    # Only proceed if courier is free and has a phone number
    if not courier or courier.status != UserStatus.FREE or not courier.phone:
        return

    for order_id, info in list(active_orders.items()):
        # Skip orders that already have a courier
        if info.get('courier_id') is not None:
            continue
        full_message = info.get('full_message')
        # Send order details and accept button to this courier
        try:
            keyboard = [[InlineKeyboardButton("✅ Qabul qilish", callback_data=f"order_{order_id}_accept")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await bot.send_message(
                chat_id=courier.id,
                text=full_message,
                reply_markup=reply_markup,
            )
            # Send location if available
            loc = info.get('location')
            if loc and isinstance(loc, dict) and 'lat' in loc and 'lng' in loc:
                try:
                    await bot.send_location(chat_id=courier.id, latitude=loc['lat'], longitude=loc['lng'])
                except Exception as e:
                    logger.error(f"Failed to send location to courier {courier.id}: {e}")
        except Exception as e:
            logger.error(f"Failed to send pending order #N{order_id} to courier {courier.id}: {e}")
            continue
        # Schedule auto assignment for this order after 15 seconds
        asyncio.create_task(auto_assign_order(order_id, full_message))

# Helpers for courier statistics
from datetime import datetime, time, timedelta
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except ImportError:
    from backports.zoneinfo import ZoneInfo  # type: ignore

async def generate_courier_report(start_dt: datetime, end_dt: datetime) -> str:
    """
    Generate a courier statistics report between start_dt and end_dt (inclusive of start_dt
    and exclusive of end_dt).  Returns the file path to the generated Excel file.  The report
    contains one row per courier with columns: No, Kuryer ismi, Buyurtmalar soni, Umumiy summa.
    Column widths are set according to the specification.
    """
    # Compute stats per courier
    stats: Dict[int, Dict[str, any]] = {}
    # Filter deliveries within the given interval (timestamps stored with tz info)
    for rec in deliveries:
        try:
            delivered_at = datetime.fromisoformat(rec.get('delivered_at'))
        except Exception:
            continue
        if delivered_at.tzinfo is None:
            # assume UTC if no timezone; convert to tz
            delivered_at = delivered_at.replace(tzinfo=ZoneInfo("UTC"))
        # check if delivered_at in [start_dt, end_dt)
        if delivered_at >= start_dt and delivered_at < end_dt:
            cid = rec.get('courier_id')
            fee = rec.get('delivery_fee', 0)
            if cid not in stats:
                stats[cid] = {'count': 0, 'sum': 0.0}
            stats[cid]['count'] += 1
            try:
                stats[cid]['sum'] += float(fee)
            except Exception:
                pass
    # Build workbook
    try:
        import openpyxl
        from openpyxl import Workbook
    except Exception as e:
        logger.error(f"openpyxl is required to generate courier report: {e}")
        raise
    wb = Workbook()
    ws = wb.active
    ws.title = "Kuryerlar"
    # Headers
    ws['A1'] = 'No'
    ws['B1'] = 'Kuryer ismi'
    ws['C1'] = 'Buyurtmalar soni'
    ws['D1'] = 'Umumiy summa'
    # Compile list of courier IDs to include (only those with any deliveries)
    # Optionally include couriers with zero; specification implies only those with deliveries for that day.
    included_couriers = [cid for cid in stats.keys()]
    # Populate rows
    row = 2
    for idx, cid in enumerate(included_couriers, start=1):
        # Find courier name
        name = next((c.name for c in couriers if c.id == cid), str(cid))
        count = stats[cid]['count']
        total = stats[cid]['sum']
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=count)
        # Format total with thousands separator and currency suffix
        ws.cell(row=row, column=4, value=f"{int(total):,} so'm")
        row += 1
    # Column width helper
    def px_to_width(px: int) -> float:
        return max((px - 5) / 7.0, 1.0)
    widths = {'A': 30, 'B': 135, 'C': 110, 'D': 110}
    for col, px in widths.items():
        try:
            ws.column_dimensions[col].width = px_to_width(px)
        except Exception:
            pass
    # Save to temporary file
    file_path = "/tmp/kuryerlar.xlsx"
    wb.save(file_path)
    return file_path

async def send_courier_report(start_dt: datetime, end_dt: datetime):
    """Generate the courier report and send it to all admins.  The file is deleted after sending."""
    file_path = await generate_courier_report(start_dt, end_dt)
    # Send to each admin
    for admin_id in [SUPER_ADMIN_ID] + list(EXTRA_ADMIN_IDS):
        try:
            with open(file_path, 'rb') as f:
                await bot.send_document(chat_id=admin_id, document=f, filename='kuryerlar.xlsx')
        except Exception as e:
            logger.error(f"Failed to send courier report to admin {admin_id}: {e}")
    # Delete the file
    try:
        os.remove(file_path)
    except Exception:
        pass

async def daily_courier_report_task():
    """
    Background task to send courier statistics report every day at 23:59 Asia/Tashkent time.
    The report covers the period from 00:00 of the current day to 23:59 (the moment of sending).
    After sending, deliveries are retained for historical purposes; filtering by time range
    ensures that only today's data is included.
    """
    tz = ZoneInfo("Asia/Tashkent")
    while True:
        now = datetime.now(tz)
        # Determine next run time at 23:59
        today = now.date()
        target = datetime.combine(today, time(23, 59), tz)
        if now >= target:
            # if already past today's target, schedule for next day
            target = datetime.combine(today + timedelta(days=1), time(23, 59), tz)
        sleep_seconds = (target - now).total_seconds()
        # Sleep until the target time
        await asyncio.sleep(sleep_seconds)
        # Determine start and end for the report (from midnight to now)
        start_dt = datetime.combine(target.date(), time(0, 0), tz)
        end_dt = datetime.combine(target.date(), time(23, 59), tz)
        # Generate and send report
        try:
            await send_courier_report(start_dt, end_dt)
        except Exception as e:
            logger.error(f"Failed to send daily courier report: {e}")

# Telegram command handlers
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = user.id
    # Ensure user is present in subscribed_users.  If new, create an entry
    global subscribed_users, phone_mapping, user_chats
    existing = next((u for u in subscribed_users if u.get('id') == user_id), None)
    if not existing:
        # Add new user with id, name and username
        new_entry = {
            'id': user_id,
            'name': user.full_name if user and user.full_name else None,
            'username': user.username if user and user.username else None,
        }
        subscribed_users.append(new_entry)
        user_chats.add(user_id)
        save_subscribed_users(subscribed_users)
    else:
        # Ensure the user's name and username are up to date
        updated = False
        if user and user.full_name and existing.get('name') != user.full_name:
            existing['name'] = user.full_name
            updated = True
        if user and user.username and existing.get('username') != user.username:
            existing['username'] = user.username
            updated = True
        if updated:
            save_subscribed_users(subscribed_users)
        user_chats.add(user_id)
    
    # Admins go directly to the admin panel
    if is_admin(user_id):
        await show_admin_panel(update, context)
        return
    # Couriers must provide a phone number before accessing their panel
    courier = get_courier(user_id)
    if courier:
        if not courier.phone:
            # Ask for a phone number
            await update.message.reply_text(
                "Iltimos to'g'ri telefon raqamingizni kiriting, mijoz sizga telefon qilishi mumkin"
            )
            context.user_data['courier_waiting_for_phone'] = True
        else:
            # Show the courier panel and offer any pending orders immediately
            await show_courier_panel(update, context)
            # If the courier is free, check for pending orders right away
            await check_pending_orders_for_courier(courier)
        return
    # Kitchen staff
    if get_kitchen(user_id):
        await show_kitchen_panel(update, context)
        return
    # Customers must provide phone number before using the panel.  If we do not already have
    # a phone entry for this user in subscribed_users, prompt them to share their phone via the
    # contact button rather than manually entering it.  We present a custom keyboard with a single
    # button that requests contact information.
    user_entry = next((u for u in subscribed_users if u.get('id') == user_id), None)
    if not user_entry or not user_entry.get('phone'):
        keyboard = [[KeyboardButton("📱 Telefon raqamni ulashish", request_contact=True)]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text(
            "Assalomu alaykum hurmatli mijoz 😊 \n\nIltimos, telefon raqamingizni yuboring. "
            "Ulashilgan raqamingiz orqali siz buyurtmangiz holatini kuzatishingiz mumkin va kerak bo'lganda kuryer siz bilan bog'lana oladi",
            reply_markup=reply_markup
        )
        # Indicate that we are waiting for the customer's contact
        context.user_data['waiting_for_phone'] = True
        return
    # Otherwise show the customer panel
    await show_customer_panel(update, context)

async def show_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Show the admin panel.  The layout is organised horizontally with two buttons per row
    for better readability.  Only the super admin sees the options to add/remove admins
    and generate the management report.
    """
    keyboard: List[List[KeyboardButton]] = []
    # First row: courier management
    keyboard.append([KeyboardButton("➕ Kuryer qo'shish"), KeyboardButton("➖ Kuryer o'chirish")])
    # Second row: branch admin management.  Rather than adding/removing branches
    # themselves, admins manage which Telegram user ID controls each branch.
    # The labels have been updated to reflect that an admin is being added or
    # removed for an existing branch.  See message handler for the workflow.
    keyboard.append([
        KeyboardButton("➕ Filial uchun admin qo'shish"),
        KeyboardButton("➖ Filial uchun admin o'chirish")
    ])
    # Third row: announcements and stats
    keyboard.append([KeyboardButton("📢 Elon yuborish"), KeyboardButton("📊 Statistika")])
    # Fourth row: courier info
    keyboard.append([KeyboardButton("🚚 Kuryer info")])
    # Super admin extras
    user_id = update.effective_user.id if update.effective_user else None
    if user_id == SUPER_ADMIN_ID:
        # Admin management buttons
        keyboard.append([KeyboardButton("➕ Admin qo'shish"), KeyboardButton("➖ Admin o'chirish")])
        # Management report
        keyboard.append([KeyboardButton("🧾 Boshqaruv")])
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "Admin paneliga xush kelibsiz!",
        reply_markup=reply_markup,
    )

async def show_courier_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    courier = get_courier(update.effective_user.id)
    status_button = "💤 Ishni tugatish" if courier.status == UserStatus.FREE else "🛵 Ishni boshlash"
    # Provide status toggle and profile button side by side
    keyboard = [[KeyboardButton(status_button), KeyboardButton("👤 Profilim")]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "Kuryer paneliga xush kelibsiz! Holatingizni o'zgartirishingiz yoki profil ma'lumotlaringizni ko'rishingiz mumkin:",
        reply_markup=reply_markup
    )

async def show_kitchen_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kitchen = get_kitchen(update.effective_user.id)
    # In the kitchen panel (branch admin), no custom keyboard should be shown
    await update.message.reply_text(
        f"Filial paneliga xush kelibsiz! {kitchen.name} filiali uchun buyurtmalar shu yerda ko'rsatiladi.",
        reply_markup=ReplyKeyboardRemove()
    )

async def show_customer_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Show the customer panel after the user has provided their phone number.  Two buttons
    are presented on the first row (Murojaat, Profilim) and a separate button for
    the guide is placed on the second row.  Note that the contact sharing button is
    not shown here; customers are asked to share their phone number via a special
    button only when first starting the bot or when updating their phone number.
    """
    # Build customer keyboard with a web app button.  The first row contains the catalog button (opening
    # https://kagoo.uz/) and the profile button.  The second row contains the contact to admin and guide.
    catalog_button = KeyboardButton(
        "🛍 KAGO Magazin",
        web_app=WebAppInfo(url="https://kagoo.uz/")
    )
    keyboard = [
        [catalog_button, KeyboardButton("👤 Profilim")],
        [KeyboardButton("✉️ Murojaat"), KeyboardButton("💵 Balansim")],
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "Hurmatli mijoz!\n\nkagoo.uz saytimiz yoki pastdagi 🛍 KAGO Magazin tugmasi orqali buyurtma berishingiz mumkin. \nIltimos buyurtma berayotganda ham botga kiritgan raqamingizdan foydalaning",
        reply_markup=reply_markup
    )

async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    # Determine phone number from contact or text.  For couriers we accept manual entry,
    # but for customers we require a contact share.  If no contact is provided and the
    # user is not a courier, prompt them to use the share button.
    courier = get_courier(user_id)
    if update.message.contact is None:
        # No contact sent
        if not courier:
            # Customer attempted to send a phone number manually; reject and remind
            await update.message.reply_text(
                "Iltimos, telefon raqamingizni ulashish tugmasi orqali yuboring.",
                reply_markup=ReplyKeyboardRemove()
            )
            return
        # For couriers we use the text field
        raw_phone = update.message.text
    else:
        # Use phone number from contact share
        raw_phone = update.message.contact.phone_number

    normalized = normalize_phone_input(raw_phone)
    if not normalized:
        await update.message.reply_text(
            "Iltimos, telefon raqamingizni to'g'ri kiriting\n(masalan: 991234567 yoki +998991234567  shaklida yozing.)",
            reply_markup=ReplyKeyboardRemove()
        )
        return

    # Update courier phone or customer phone in subscribed_users and phone_mapping
    # Determine if this user is known in subscribed_users
    global subscribed_users, phone_mapping
    entry = next((u for u in subscribed_users if u.get('id') == user_id), None)
    if courier:
        # Save phone into the courier object
        courier.phone = normalized
        save_data(COURIERS_FILE, [c.dict() for c in couriers])
        # Update subscribed_users entry as well
        if entry:
            entry['phone'] = normalized
            # Optionally store name/username if not already present
            if not entry.get('name') and update.effective_user and update.effective_user.full_name:
                entry['name'] = update.effective_user.full_name
            if not entry.get('username') and update.effective_user and update.effective_user.username:
                entry['username'] = update.effective_user.username
        else:
            subscribed_users.append({'id': user_id, 'phone': normalized})
        # Update phone_mapping (remove any old mapping for this user)
        to_remove = [p for p, uid in phone_mapping.items() if uid == user_id]
        for p in to_remove:
            phone_mapping.pop(p, None)
        phone_mapping[normalized] = user_id
        save_subscribed_users(subscribed_users)
        # Clear flags
        context.user_data.pop('courier_waiting_for_phone', None)
        context.user_data.pop('changing_phone', None)
        await update.message.reply_text(
            f"Sizning {normalized} raqamingizni eslab qoldik!",
            reply_markup=ReplyKeyboardRemove()
        )
        # After providing or updating phone, offer pending orders if any
        await check_pending_orders_for_courier(courier)
        await show_courier_panel(update, context)
        return

    # Customer phone: either new registration or phone change
    # If the user is changing their phone (triggered by inline button), remove any existing mapping
    if context.user_data.get('changing_phone'):
        to_remove = [p for p, uid in phone_mapping.items() if uid == user_id]
        for p in to_remove:
            phone_mapping.pop(p, None)
        context.user_data['changing_phone'] = False
    # Update subscribed_users entry
    if entry:
        entry['phone'] = normalized
        # update name/username if available
        if update.effective_user:
            if update.effective_user.full_name:
                entry['name'] = update.effective_user.full_name
            if update.effective_user.username:
                entry['username'] = update.effective_user.username
    else:
        # Add a new entry for this user
        entry = {
            'id': user_id,
            'phone': normalized,
            'name': update.effective_user.full_name if update.effective_user else None,
            'username': update.effective_user.username if update.effective_user else None,
        }
        subscribed_users.append(entry)
        user_chats.add(user_id)
    # Update phone_mapping
    phone_mapping[normalized] = user_id
    save_subscribed_users(subscribed_users)
    # Clear waiting flag
    context.user_data.pop('waiting_for_phone', None)
    await update.message.reply_text(
        f"Sizning {normalized} raqamingizni eslab qoldik!",
        reply_markup=ReplyKeyboardRemove()
    )
    # Show the customer panel
    await show_customer_panel(update, context)

async def handle_customer_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    # Normalise text (may be None for non‑text messages)
    text = update.message.text if update.message.text else ""

    # If we are waiting for the customer's phone number, only accept via contact share.
    # If the customer tries to type the phone manually, remind them to use the share button.
    if context.user_data.get('waiting_for_phone'):
        await update.message.reply_text(
            "Iltimos, telefon raqamingizni ulashish tugmasi orqali yuboring."
        )
        return

    # If the user is changing their phone number, they must share the new number via contact.  Reject manual entry.
    if context.user_data.get('changing_phone'):
        await update.message.reply_text(
            "Iltimos, yangi telefon raqamingizni ulashish tugmasi orqali yuboring."
        )
        return

    # Handle help/guide button
    if text == "💵 Balansim":
        # Compose a guide message.  If the user has a registered phone
        # number, fetch their current cashback balance and promo code from
        # the server.  If no phone is registered, prompt them to share
        # their phone via the profile section.
        user = update.effective_user
        user_id = user.id if user else None
        phone = None
        for u in subscribed_users:
            if u.get('id') == user_id and u.get('phone'):
                phone = u.get('phone')
                break
        message = (
            "💵 Balansim\n\n"
        )
        if phone:
            try:
                resp = requests.get(f"{SERVER_API_BASE_URL}/api/cashback/{phone}")
                if resp.status_code == 200:
                    data = resp.json()
                    bal = data.get('cashback', 0)
                    promo = data.get('promoCode') or data.get('promo')
                    message += (
                        f"💵 Balansingiz: {bal:,.0f} so'm.\n"
                    )
                    if promo:
                        message += (
                            f"🔢 Keshbek promo kodingiz: {promo}. "
                            "\n\nℹ️ Ushbu keshbek promo kodni boshqa promo kodlar bilan bir vaqtda ishlatishingiz mumkin, promo kodingizni hech kimga bermang!\nHar bir xaridingiz uchun 1% keshbek yig'ilib boradi va keshbek miqdori 10 ming so'mdan oshganida siz undan foydalana olasiz\n\nKeshbek promo kodni mahsulot buyurtma berayotganda Promo kod kiritish maydoniga kiriting va narx avtomatik kamayadi"
                        )
                else:
                    # If the request fails, silently ignore and just show the generic guide
                    pass
            except Exception as e:
                # Log but don't expose exception details to the user
                logging.error(f"Failed to fetch cashback for phone {phone}: {e}")
        else:
            message += (
                "Telefon raqamingiz botda ro'yxatdan o'tmagan. "
                "Profil bo'limiga o'tib, raqamingizni ulashish tugmasi orqali kiritishingiz mumkin.\n"
            )
        # Final note
        message += " "
        await update.message.reply_text(message)
        return

    # Handle contact button
    if text == "✉️ Murojaat":
        await update.message.reply_text(
            "Iltimos, murojaatingizni yuboring. Admin tez orada javob beradi."
        )
        context.user_data['waiting_for_message'] = True
        return

    # Handle profile button
    if text == "👤 Profilim":
        user = update.effective_user
        # Retrieve phone from subscribed_users
        phone = next((u.get('phone') for u in subscribed_users if u.get('id') == user_id), "Noma'lum")
        # Inline button to change phone
        keyboard = [[InlineKeyboardButton("📱 Raqamni almashtirish", callback_data="change_phone")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        username_disp = f"@{user.username}" if user and user.username else "Nomalum"
        await update.message.reply_text(
            f"👤 Profil ma'lumotlari:\n"
            f"🆔 ID: {user_id}\n"
            f"👤 Ism: {user.full_name}\n"
            f"📱 Telefon: {phone}\n"
            f"🔗 Username: {username_disp}\n",
            reply_markup=reply_markup
        )
        return

    # If the user is submitting a message to the admin (any type)
    if context.user_data.get('waiting_for_message'):
        # Forward the actual message (photo, audio, etc.) to all admins preserving its type
        # We no longer persist messages on disk.
        for admin_id in [SUPER_ADMIN_ID] + list(EXTRA_ADMIN_IDS):
            try:
                await update.message.copy(chat_id=admin_id)
                # Send additional info with a reply button
                keyboard = [[InlineKeyboardButton("✍️ Javob yozish", callback_data=f"reply_{user_id}")]]
                reply_markup = InlineKeyboardMarkup(keyboard)
                # Determine phone from subscribed_users
                phone = next((u.get('phone') for u in subscribed_users if u.get('id') == user_id), 'Nomalum')
                await bot.send_message(
                    chat_id=admin_id,
                    text=(
                        f"📩 Yangi murojaat:\n"
                        f"👤 Foydalanuvchi: {update.effective_user.full_name}\n"
                        f"🆔 ID: {user_id}\n"
                        f"📱 Telefon: {phone}"
                    ),
                    reply_markup=reply_markup
                )
            except Exception as e:
                logger.error(f"Failed to forward message to admin {admin_id}: {e}")
        await update.message.reply_text("Xabaringiz adminga yuborildi. Tez orada javob beriladi.")
        context.user_data['waiting_for_message'] = False
        return

    # Unknown command
    await update.message.reply_text("Noma'lum buyruq. Iltimos, menyudan foydalaning.")

async def handle_courier_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text if update.message.text else ""
    courier = get_courier(update.effective_user.id)
    # If courier is being asked to provide a phone number initially
    if context.user_data.get('courier_waiting_for_phone'):
        normalized = normalize_phone_input(text)
        if not normalized:
            await update.message.reply_text("Iltimos, telefon raqamingizni to'g'ri kiriting\n(masalan: 991234567 yoki +998991234567  shaklida yozing.)")
            return
        courier.phone = normalized
        save_data(COURIERS_FILE, [c.dict() for c in couriers])
        context.user_data['courier_waiting_for_phone'] = False
        await update.message.reply_text(f"Sizning {normalized} raqamingizni eslab qoldik!")
        # After providing phone number, courier may now be available for pending orders
        await check_pending_orders_for_courier(courier)
        await show_courier_panel(update, context)
        return
    # If courier is changing phone number
    if context.user_data.get('changing_phone'):
        normalized = normalize_phone_input(text)
        if not normalized:
            await update.message.reply_text("Iltimos, telefon raqamingizni to'g'ri kiriting\n(masalan: 991234567 yoki +998991234567  shaklida yozing.)")
            return
        courier.phone = normalized
        save_data(COURIERS_FILE, [c.dict() for c in couriers])
        context.user_data['changing_phone'] = False
        await update.message.reply_text(f"Raqamingiz yangilandi: {normalized}")
        # After phone change, courier may now be available for pending orders
        await check_pending_orders_for_courier(courier)
        # Show profile again after change
        await show_courier_panel(update, context)
        return
    # Handle profile viewing
    if text == "👤 Profilim":
        user = update.effective_user
        phone = courier.phone if courier.phone else "Nomalum"
        keyboard = [[InlineKeyboardButton("📱 Raqamni almashtirish", callback_data="change_phone")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            f"👤 Profil ma'lumotlari:\n"
            f"🆔 ID: {courier.id}\n"
            f"👤 Ism: {courier.name}\n"
            f"📱 Telefon: {phone}\n"
            f"🔗 Username: @{user.username if user.username else 'Nomalum'}\n",
            reply_markup=reply_markup
        )
        return
    # Handle status toggling
    if text == "🛵 Ishni boshlash":
        courier.status = UserStatus.FREE
        save_data(COURIERS_FILE, [c.dict() for c in couriers])
        await update.message.reply_text(
            "Xush kelibsiz hurmatli kuryer! 😎 \n\nSiz ishni boshladingiz, siz endi yangi buyurtmalarni qabul qilasiz ✅ \nYo'llarda ehtiyot bo'ling, Omad! 💪"
        )
        # After becoming free, offer any pending orders to this courier
        await check_pending_orders_for_courier(courier)
        await show_courier_panel(update, context)
        return
    if text == "💤 Ishni tugatish":
        courier.status = UserStatus.UNAVAILABLE
        save_data(COURIERS_FILE, [c.dict() for c in couriers])
        await update.message.reply_text("🛑 Siz dam olish rejimiga o'tdingiz! Endi sizni yangi buyurtmalar bezovta qilmaydi! \n\nYaxshi dam olib qayting!🛌")
        await show_courier_panel(update, context)
        return
    # Any other text message is considered unknown in the courier panel
    await update.message.reply_text("Noma'lum buyruq. Iltimos, menyudan foydalaning.")

async def handle_callback_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    user_id = query.from_user.id
    
    await query.answer()
    
    if data == "add_courier":
        context.user_data['adding_courier'] = True
        context.user_data['adding_courier_name'] = False
        await query.edit_message_text("Kuryer ID raqamini yuboring:")
    
    elif data == "add_kitchen":
        # Initiate branch admin assignment via inline keyboard.  Ask for branch name first.
        context.user_data['adding_branch_admin_name'] = True
        context.user_data['adding_branch_admin_id'] = False
        context.user_data.pop('branch_name', None)
        await query.edit_message_text(
            "Filial nomini yuboring\n\n⚠️ Diqqat filial nomini saytdagi bilan bir xil qilib kiriting! Harflarning katta-kichikligi, bo'sh joy va h.k. saytdagi bilan bir xil nom bo'lishi shart!"
        )
    
    elif data == "remove_courier":
        context.user_data['removing_courier'] = True
        await query.edit_message_text("O'chiriladigan kuryer ID raqamini yuboring:")
    
    elif data == "remove_kitchen":
        # Initiate branch admin removal via inline keyboard.  Ask for branch name.
        context.user_data['removing_branch_admin'] = True
        await query.edit_message_text(
            "O'chiriladigan filial nomini yuboring (saytdagi bilan bir xil)."
        )
    
    elif data == "send_announcement":
        context.user_data['waiting_for_announcement'] = True
        await query.edit_message_text("📢 Elonni yuboring:")
    
    elif data == "show_stats":
        await query.edit_message_text(f"📊 Statistika:\n👥 Foydalanuvchilar soni: {len(user_chats)}\n🚴 Kuryerlar soni: {len(couriers)}\n🏪 Filiallar soni: {len(kitchens)}")
    
    elif data.startswith("reply_"):
        target_user = int(data[6:])
        context.user_data['replying_to'] = target_user
        context.user_data['waiting_for_reply'] = True
        await query.edit_message_text("✍️ Javobingizni yozing:")
    
    elif data == "change_phone":
        # Both customers and couriers can change their phone number.  Set a flag and prompt accordingly.
        context.user_data['changing_phone'] = True
        user_id_cb = query.from_user.id
        # Determine if this user is a courier
        courier_cb = get_courier(user_id_cb)
        if courier_cb:
            # For couriers we allow manual entry; edit the inline message to prompt
            await query.edit_message_text("Yangi telefon raqamingizni yuboring:")
        else:
            # For customers, we require contact sharing.  Since inline keyboards cannot request contact,
            # send a new message with a reply keyboard that requests contact.  Also edit the original message.
            await query.edit_message_text("Yangi telefon raqamingizni ulashishingiz mumkin.")
            keyboard = [[KeyboardButton("📱 Telefon raqamni ulashish", request_contact=True)]]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
            try:
                await bot.send_message(
                    chat_id=user_id_cb,
                    text="Iltimos, yangi telefon raqamingizni ulashish tugmasi orqali yuboring.",
                    reply_markup=reply_markup
                )
            except Exception as e:
                logger.error(f"Failed to prompt customer {user_id_cb} for phone change: {e}")
    
    elif data.startswith("order_"):
        await handle_order_actions(query, data[6:])
    
    else:
        await query.edit_message_text("Noma'lum buyruq!")

async def handle_order_actions(query, action):
    courier = get_courier(query.from_user.id)
    if not courier:
        # Only couriers can perform these actions
        await query.edit_message_text("Siz kuryer emassiz!")
        return

    # Attempt to parse the order ID from the callback data. If the callback data is of the form
    # "<order_id>_<subaction>", split it accordingly; otherwise fall back to the courier's current order.
    order_id: Optional[int] = None
    subaction: str = action
    if '_' in action:
        maybe_id, maybe_action = action.split('_', 1)
        if maybe_id.isdigit():
            order_id = int(maybe_id)
            subaction = maybe_action
    if order_id is None:
        order_id = courier.current_order
    
    # Validate that this order is being tracked
    if order_id is None or order_id not in active_orders:
        await query.edit_message_text("Bu buyurtma haqida ma'lumot topilmadi.")
        return
    
    # Convenience references
    order_info = active_orders[order_id]
    assigned_courier_id = order_info.get('courier_id')
    customer_id = order_info.get('customer_id')
    location = order_info.get('location')

    # Accepting the order
    if subaction == "accept":
        # If the order is already assigned to another courier, prevent further acceptance
        if assigned_courier_id is not None and assigned_courier_id != courier.id:
            await query.edit_message_text("Bu buyurtma allaqachon boshqa kuryerga biriktirilgan!")
            return
        # Assign the order to this courier
        order_info['courier_id'] = courier.id
        courier.status = UserStatus.BUSY
        courier.current_order = order_id
        save_data(COURIERS_FILE, [c.dict() for c in couriers])

        # Notify admins that the courier accepted
        await send_to_admins(f"🚴 Kuryer {courier.name} #N{order_id} - buyurtmani qabul qildi!")

        # Notify the customer that the order has been accepted and provide courier details
        if customer_id:
            try:
                # Build detailed message
                msg = "Sizning buyurtmangiz qabul qilindi\n"
                # Courier name
                msg += f"\nKuryer ismi: {courier.name}"
                # Courier phone
                if courier.phone:
                    msg += f"\nKuryer telefon raqami: {courier.phone}"
                # Courier username
                if query.from_user and query.from_user.username:
                    msg += f"\nKuryer: @{query.from_user.username}"
                await bot.send_message(chat_id=customer_id, text=msg)
            except Exception as e:
                logger.error(f"Failed to send acceptance message to customer {customer_id}: {e}")

        # Edit the original message to acknowledge acceptance
        await query.edit_message_text("✅ Buyurtmani qabul qildingiz!")

        # Resend full order details to the courier for reference
        full_msg = order_info.get('full_message')
        if full_msg:
            try:
                # Send without specifying parse_mode to avoid Markdown parsing errors
                await query.message.reply_text(
                    full_msg,
                )
            except Exception as e:
                logger.error(f"Failed to resend order details to courier {courier.id}: {e}")

        # Send a new message with instructions and the next step button
        keyboard = [[InlineKeyboardButton("📦 Buyurtma menda", callback_data=f"order_{order_id}_received")]]
        await query.message.reply_text(
            "Barchasini olgandan keyin pastdagi 'Buyurtma menda' tugmasini bosing!",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return

    # Confirm that the courier has picked up the order
    if subaction == "received":
        # Only the assigned courier can proceed
        if assigned_courier_id != courier.id:
            await query.edit_message_text("Bu buyurtma sizga biriktirilmagan.")
            return
        # Notify customer that courier has picked up the order
        if customer_id:
            try:
                await bot.send_message(chat_id=customer_id, text="📦 Kuryer buyurtmangizni qabul qilib oldi")
            except Exception as e:
                logger.error(f"Failed to send pickup message to customer {customer_id}: {e}")

        # Edit message and send next button
        await query.edit_message_text("✅ Buyurtma sizda ekanligi tasdiqlandi!")
        keyboard = [[InlineKeyboardButton("🏠 Mijoz tomon", callback_data=f"order_{order_id}_delivering")]]
        await query.message.reply_text(
            "Barcha buyurtmalarni olgan bo'lsangiz pastdagi 'Mijoz tomon' tugmasini bosing!",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return

    # Courier is heading to the customer
    if subaction == "delivering":
        if assigned_courier_id != courier.id:
            await query.edit_message_text("Bu buyurtma sizga biriktirilmagan.")
            return
        # Notify customer that courier is heading to the customer
        if customer_id:
            try:
                await bot.send_message(chat_id=customer_id, text="🚗 Kuryer siz tomonga qarab jo'nadi")
            except Exception as e:
                logger.error(f"Failed to send en route message to customer {customer_id}: {e}")
        # Send the location to the courier if available
        if location and isinstance(location, dict) and 'lat' in location and 'lng' in location:
            try:
                await bot.send_location(chat_id=courier.id, latitude=location['lat'], longitude=location['lng'])
            except Exception as e:
                logger.error(f"Failed to send location to courier {courier.id}: {e}")
        
        await query.edit_message_text("🚚 Mijoz tomon yo'l oldingiz!")
        keyboard = [[InlineKeyboardButton("✅ Keldim", callback_data=f"order_{order_id}_arrived")]]
        await query.message.reply_text(
            "Endi yuqoridagi mijoz manziliga boring. Yetib borgach pastdagi 'Keldim' tugmasini bosing!",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return

    # Courier has arrived at the customer's location
    if subaction == "arrived":
        if assigned_courier_id != courier.id:
            await query.edit_message_text("Bu buyurtma sizga biriktirilmagan.")
            return
        # Notify customer that courier has arrived
        if customer_id:
            try:
                await bot.send_message(
                    chat_id=customer_id,
                    text="🚘 Kuryer yetib keldi! Iltimos, buyurtmangizni qabul qilib kuryerimizga to'lov qiling!"
                )
            except Exception as e:
                logger.error(f"Failed to send arrival message to customer {customer_id}: {e}")
        await query.edit_message_text("✅ Siz mijozga yetib kelganingiz tasdiqlandi!")
        keyboard = [[InlineKeyboardButton("✅ Bajarildi", callback_data=f"order_{order_id}_completed")]]
        await query.message.reply_text(
            "Buyurtmani topshirib to'lovni qabul qilgach pastdagi 'Bajarildi' tugmasini bosing!",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return

    # Completion of the order
    if subaction == "completed":
        if assigned_courier_id != courier.id:
            await query.edit_message_text("Bu buyurtma sizga biriktirilmagan.")
            return
        # Release the courier
        courier.status = UserStatus.FREE
        courier.current_order = None
        save_data(COURIERS_FILE, [c.dict() for c in couriers])
        # Notify customer that order is completed and award cashback on the dishes subtotal
        cashback_earned = 0
        # Retrieve order info from active_orders so we can award cashback. This lookup must occur
        # before the order is removed from active_orders.
        order_info = active_orders.get(order_id)
        if order_info:
            phone = order_info.get('phone')
            dishes_total = order_info.get('dishes_total')
            try:
                if phone and dishes_total:
                    # Award 1% cashback on the dishes subtotal
                    cashback_earned = int(dishes_total * 0.01)
                    if cashback_earned > 0:
                        cashbacks_raw = load_data('cashbacks.json', dict)
                        # Normalise existing entries: either a number or an object
                        cashbacks = {}
                        for k, v in cashbacks_raw.items():
                            if isinstance(v, (int, float)):
                                cashbacks[k] = {'balance': v}
                            elif isinstance(v, dict):
                                bal = v.get('balance') if isinstance(v.get('balance'), (int, float)) else v.get('cashback', 0)
                                code = v.get('promo') or v.get('code')
                                cashbacks[k] = {'balance': bal or 0, 'promo': code}
                        entry = cashbacks.get(phone, {'balance': 0})
                        # Generate a unique six‑character alphanumeric promo code if one
                        # does not exist.  Codes consist of uppercase letters and digits
                        # (e.g. "A23BU7").  If a generated code already exists, try
                        # again until a unique one is produced.
                        if not entry.get('promo'):
                            existing_codes = {e.get('promo') for e in cashbacks.values() if e.get('promo')}
                            code = None
                            chars = string.ascii_uppercase + string.digits
                            while True:
                                code_candidate = ''.join(random.choice(chars) for _ in range(6))
                                if code_candidate not in existing_codes:
                                    code = code_candidate
                                    break
                            entry['promo'] = code
                        # Update the balance
                        entry['balance'] = entry.get('balance', 0) + cashback_earned
                        cashbacks[phone] = entry
                        save_data('cashbacks.json', cashbacks)
            except Exception as e:
                logger.error(f"Failed to update cashback for {phone}: {e}")
        # Prepare a completion message for the customer. Include cashback information
        # only if some was earned.
        if customer_id:
            try:
                message = "Xizmatimizdan foydalanganingiz uchun tashakkur!\nSizni yana kutib qolamiz! ♥️"
                if cashback_earned > 0 and phone:
                    # Reload cashbacks to fetch the latest balance and promo code
                    cb_raw = load_data('cashbacks.json', dict)
                    # Normalise again to ensure consistent structure
                    cb_entry = None
                    if isinstance(cb_raw.get(phone), (int, float)):
                        cb_entry = {'balance': cb_raw.get(phone), 'promo': None}
                    elif isinstance(cb_raw.get(phone), dict):
                        bal = cb_raw[phone].get('balance') if isinstance(cb_raw[phone].get('balance'), (int, float)) else cb_raw[phone].get('cashback', 0)
                        code = cb_raw[phone].get('promo') or cb_raw[phone].get('code')
                        cb_entry = {'balance': bal or 0, 'promo': code}
                    if cb_entry:
                        message += f"\n\nSizga xaridingiz uchun {cashback_earned:,.0f} so'm keshbek berildi! 🎉"
                        if cb_entry.get('promo'):
                            message += f"\n\nKeshbekdan foydalanish uchun 💵 Balansim tugmasini bosing"
                await bot.send_message(
                    chat_id=customer_id,
                    text=message
                )
            except Exception as e:
                logger.error(f"Failed to send completion message to customer {customer_id}: {e}")
        # Inform admin
        await send_to_admins(f"✅ Kuryer {courier.name} #N{order_id} - buyurtmani yetkazib berdi!")
        # Edit message to confirm completion
        await query.edit_message_text("✅ Buyurtma muvaffaqiyatli yakunlandi!")
        # Before removing the order from active orders, record this delivery for statistics
        # Retrieve delivery fee from active_orders (if present) and record completion time in Asia/Tashkent
        info = active_orders.get(order_id)
        if info:
            try:
                from zoneinfo import ZoneInfo as _Z
            except ImportError:
                from backports.zoneinfo import ZoneInfo as _Z  # type: ignore
            tz = _Z("Asia/Tashkent")
            delivered_at = datetime.now(tz).isoformat()
            delivery_fee = info.get('delivery_fee', 0)
            deliveries.append({
                'order_id': order_id,
                'courier_id': courier.id,
                'delivered_at': delivered_at,
                'delivery_fee': delivery_fee,
            })
            # Persist deliveries to file
            save_data(DELIVERIES_FILE, deliveries)
        # Remove order from active orders
        active_orders.pop(order_id, None)
        # After finishing the order, the courier becomes free; offer any pending orders
        await check_pending_orders_for_courier(courier)
        return

# FastAPI endpoint
app = FastAPI()

@app.post("/api/orders")
async def receive_order_notification(
    request: Request, 
    authorization: str = Header(None)
):
    logger.info("Received order notification request")
    
    if authorization != f"Bearer {FASTAPI_SECRET_TOKEN}":
        logger.error("Invalid auth token received")
        raise HTTPException(status_code=401, detail="Invalid auth token")
    
    try:
        order_data = await request.json()
        order = OrderModel(**order_data)
        
        # Prepare order message
        items_list = []
        for item in order.items:
            item_str = f"\n🏪 {item.restaurant}\n- {item.name} x{item.quantity} ({item.price_per_unit:,.0f} so'm) = {item.total_price:,.0f} so'm"
            if item.container:
                item_str += f"\n  - {item.container.name} x{item.container.quantity} ({item.container.price_per_unit:,.0f} so'm) = {item.container.total_price:,.0f} so'm"
            items_list.append(item_str)
        
        items_text = "\n".join(items_list)
        
        # We'll send a Telegram location separately; no need to embed a Google Maps link in the message
        location_info = ""
        lat = None
        lng = None
        if order.customer.location and 'lat' in order.customer.location and 'lng' in order.customer.location:
            lat = order.customer.location['lat']
            lng = order.customer.location['lng']
        
        # Determine username for the customer from subscribed_users.  We avoid
        # making API calls here because they can fail and prevent the order message
        # from being sent.  We match the customer phone to an entry in the
        # subscribed_users data and extract the stored username.
        username_display = "aniqlanmadi"
        # Normalize the phone from the order (ignore non‑digits and use last 9 digits)
        def _extract_digits(p: str) -> str:
            digits = ''.join(filter(str.isdigit, p)) if p else ''
            if digits.startswith('998'):
                digits = digits[3:]
            if len(digits) >= 9:
                return digits[-9:]
            return digits
        order_phone_digits = _extract_digits(order.customer.phone)
        matched_entry = None
        for u in subscribed_users:
            phone_u = u.get('phone')
            if not phone_u:
                continue
            if _extract_digits(phone_u) == order_phone_digits:
                matched_entry = u
                break
        if matched_entry:
            uname = matched_entry.get('username')
            if uname:
                username_display = f"@{uname}"

        full_message = (
            f"🛒 Yangi buyurtma (#N{order.order_id})\n"
            f"{items_text}\n\n"
            f"🧾 Mahsulotlar narxi: {order.totals.dishes:,.0f} so'm\n"
            f"🫙 Idishlar narxi: {order.totals.containers:,.0f} so'm\n"
            f"🚚 Yetkazib berish: {order.totals.delivery:,.0f} so'm\n"
            f"🎟 Chegirma: {order.totals.discount:,.0f} so'm\n"
            f"💳 To'lov usuli: {'Naqd pul' if order.payment_method == 'cash' else 'Karta orqali'}\n\n"
            f"💰 Yakuniy summa: {order.totals.final:,.0f} so'm\n"
            + (f"\n📝 Kuryer uchun izoh: {order.customer.note}\n" if order.customer.note else "")
            + f"\n👤 Mijoz:\n"
            f"🪪 Ism: {order.customer.name}\n"
            f"📱 Telefon: {order.customer.phone}\n"
            f"🔗 Username: {username_display}\n\n"
            f"⌚️ Buyurtma vaqti: {order.created_at}"
        )
        
        # Send full order to admins without specifying parse_mode to avoid Markdown parsing errors
        await send_to_admins(full_message)
        # Send location separately to admins if available
        if lat is not None and lng is not None:
            for admin_id in [SUPER_ADMIN_ID] + list(EXTRA_ADMIN_IDS):
                try:
                    await bot.send_location(chat_id=admin_id, latitude=lat, longitude=lng)
                except Exception as e:
                    logger.error(f"Failed to send location to admin {admin_id}: {e}")
        
        # Determine if the customer has a matching Telegram user. We perform this lookup once here
        customer_id = find_user_by_phone(order.customer.phone)

        # Register the order in our active_orders mapping
        # We store the customer_id (if found) and leave courier_id unset until accepted or auto‑assigned.
        # In addition, persist the customer's phone number and the dishes subtotal so that
        # cashback can be awarded when the order is completed. We also record the
        # delivery fee and order creation timestamp for later statistics.
        active_orders[order.order_id] = {
            'customer_id': customer_id,
            'courier_id': None,
            'location': order.customer.location if order.customer.location else None,
            'full_message': full_message,
            'delivery_fee': order.totals.delivery,
            'created_at': order.created_at,
            'phone': order.customer.phone,
            'dishes_total': order.totals.dishes,
        }

        # Determine available couriers at this time
        available_couriers = get_available_couriers()
        if available_couriers:
            # Send full order to all currently available couriers with an accept button
            keyboard = [[InlineKeyboardButton("✅ Qabul qilish", callback_data=f"order_{order.order_id}_accept")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            for courier in available_couriers:
                try:
                    await bot.send_message(
                        chat_id=courier.id,
                        text=full_message,
                        reply_markup=reply_markup,
                    )
                    # Send location separately to the courier
                    if lat is not None and lng is not None:
                        try:
                            await bot.send_location(chat_id=courier.id, latitude=lat, longitude=lng)
                        except Exception as e:
                            logger.error(f"Failed to send location to courier {courier.id}: {e}")
                except Exception as e:
                    logger.error(f"Failed to send to courier {courier.id}: {e}")

            # Schedule an automatic assignment if no courier accepts within the timeout
            asyncio.create_task(auto_assign_order(order.order_id, full_message))
        else:
            # No available couriers: notify admins that the order is in waiting mode
            try:
                await send_to_admins(
                    f"Barcha kuryerlar band #N{order.order_id} - buyurtma kutish rejimiga o'tkazildi. "
                    "Kuryerlardan biri ishni yakunlaganda yoki yangi kuryer qo'shilsa buyurtma avtomatik yo'naltiriladi"
                )
            except Exception:
                pass
            # Do not schedule auto assignment yet.  The order will be sent to the next available courier
            # when one becomes free or is added.  The auto_assign_task will be scheduled at that time.
        
        # Send specific items to each kitchen
        restaurant_items = {}
        for item in order.items:
            if item.restaurant not in restaurant_items:
                restaurant_items[item.restaurant] = []
            item_str = f"- {item.name} x{item.quantity} ({item.price_per_unit:,.0f} so'm) = {item.total_price:,.0f} so'm"
            if item.container:
                item_str += f"\n  - {item.container.name} x{item.container.quantity} ({item.container.price_per_unit:,.0f} so'm) = {item.container.total_price:,.0f} so'm"
            restaurant_items[item.restaurant].append(item_str)
        
        for restaurant, items in restaurant_items.items():
            kitchen_message = f"🛒 Yangi buyurtma (#N{order.order_id})\n" + "\n".join(items)
            await send_to_kitchen(restaurant, kitchen_message)
        
        # Send notification to customer
        customer_id = find_user_by_phone(order.customer.phone)
        if customer_id:
            await bot.send_message(
                chat_id=customer_id,
                text=full_message,
            )
        
        return {"success": True}
        
    except Exception as e:
        logger.error(f"Error processing order: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=str(e))

async def auto_assign_order(order_id: int, message: str):
    # Wait for a short period for couriers to manually accept. User specification: 15 seconds
    await asyncio.sleep(15)

    # If a courier has already been assigned (either manually or previously), do nothing
    current_courier = get_courier_by_order(order_id)
    if current_courier is not None:
        return

    # Otherwise pick a random available courier to assign the order
    available = get_available_couriers()
    if not available:
        logger.warning(f"No available couriers to auto-assign order #N{order_id}")
        return

    courier = random.choice(available)
    # Set courier status and assign order id
    courier.status = UserStatus.BUSY
    courier.current_order = order_id
    save_data(COURIERS_FILE, [c.dict() for c in couriers])

    # Update active_orders mapping
    if order_id in active_orders:
        active_orders[order_id]['courier_id'] = courier.id

    # Send message to the assigned courier with a "Buyurtma menda" button
    keyboard = [[InlineKeyboardButton("📦 Buyurtma menda", callback_data=f"order_{order_id}_received")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        await bot.send_message(
            chat_id=courier.id,
            text=(
                f"✅ Sizga avtomatik ravishda buyurtma biriktirildi:\n\n{message}\n\n"
                "Barchasini olgandan keyin pastdagi 'Buyurtma menda' tugmasini bosing!"
            ),
            reply_markup=reply_markup,
        )
    except Exception as e:
        logger.error(f"Failed to notify courier {courier.id} about auto-assigned order #N{order_id}: {e}")

    # Notify admins about auto-assignment
    await send_to_admins(f"🚴 Kuryer {courier.name} ga #N{order_id} - buyurtma avtomatik biriktirildi!")

    # Notify the customer that the order has been accepted automatically
    if order_id in active_orders:
        customer_id = active_orders[order_id].get('customer_id')
        if customer_id:
            try:
                # Construct acceptance message with courier details
                msg = "Sizning buyurtmangiz qabul qilindi\n"
                msg += f"\nKuryer ismi: {courier.name}"
                if courier.phone:
                    msg += f"\nKuryer telefon raqami: {courier.phone}"
                # Attempt to get the courier's username
                try:
                    chat_info = await bot.get_chat(courier.id)
                    if chat_info.username:
                        msg += f"\nKuryer: @{chat_info.username}"
                except Exception:
                    pass
                await bot.send_message(chat_id=customer_id, text=msg)
            except Exception as e:
                logger.error(f"Failed to send acceptance message to customer {customer_id}: {e}")

# Bot setup
async def setup_bot():
    await application.initialize()
    await application.start()
    await application.updater.start_polling()
    logger.info("Bot polling started")

@app.on_event("startup")
async def on_startup():
    # Command handlers
    application.add_handler(CommandHandler("start", start))
    # Register contact handler before the general message handler so that contact
    # messages are not caught by the catch‑all handler.  The general handler uses
    # a broad filter to accept any non‑command message, enabling forwarding of
    # photos, audio, video and other media types.
    application.add_handler(MessageHandler(filters.CONTACT, handle_contact))
    application.add_handler(MessageHandler(~filters.COMMAND, handle_message))
    # Callback handlers
    application.add_handler(CallbackQueryHandler(handle_callback_query))
    
    # Start bot
    asyncio.create_task(setup_bot())
    # Start daily courier report scheduler
    asyncio.create_task(daily_courier_report_task())
    logger.info("Service started")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if is_admin(user_id):
        await handle_admin_message(update, context)
    elif get_courier(user_id):
        await handle_courier_message(update, context)
    elif get_kitchen(user_id):
        pass  # Kitchen staff don't need special message handling
    else:
        await handle_customer_message(update, context)

async def handle_admin_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text
    
    # Contextdagi holatlarni tekshiramiz
    if 'adding_courier' in context.user_data and context.user_data['adding_courier']:
        try:
            courier_id = int(text)
            context.user_data['courier_id'] = courier_id
            await update.message.reply_text("Kuryer ismini yuboring:")
            context.user_data['adding_courier'] = False
            context.user_data['adding_courier_name'] = True
            return
        except ValueError:
            await update.message.reply_text("Noto'g'ri ID format! Faqat raqam yuboring.")
            return
    
    elif 'adding_courier_name' in context.user_data and context.user_data['adding_courier_name']:
        courier_id = context.user_data['courier_id']
        couriers.append(CourierModel(id=courier_id, name=text))
        save_data(COURIERS_FILE, [c.dict() for c in couriers])
        
        await update.message.reply_text(f"✅ Kuryer {text} (ID: {courier_id}) qo'shildi!")
        context.user_data['adding_courier_name'] = False
        await show_admin_panel(update, context)
        return
    
    elif context.user_data.get('adding_branch_admin_name'):
        # Stage 1: The admin has provided a branch name.  Validate against the
        # list of branch names defined on the website (restaurants.json).  If
        # the branch does not exist, inform the admin and abort the operation.
        branch_name_input = text.strip()
        def load_site_branches() -> List[str]:
            """
            Fetch the list of branch names from the server's /api/restaurants endpoint.
            If the API call fails or returns invalid data, fall back to reading
            from the local restaurants.json file.  Returns a list of trimmed
            restaurant names (strings).
            """
            names: List[str] = []
            # First attempt to read from the server API
            try:
                url = f"{SERVER_API_BASE_URL}/api/restaurants"
                response = requests.get(url, timeout=5)
                # Ensure we got a successful response with JSON
                if response.status_code == 200:
                    data = response.json()
                    if isinstance(data, list):
                        for item in data:
                            name = item.get('name') if isinstance(item, dict) else None
                            if name and isinstance(name, str):
                                names.append(name.strip())
                        return names
            except Exception as e:
                # Log API-related errors but continue to fallback
                logger.error(f"Error fetching restaurants from API: {e}")
            # If API call fails or data is malformed, try reading from local file
            try:
                if os.path.exists('restaurants.json'):
                    with open('restaurants.json', 'r') as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            for item in data:
                                name = item.get('name') if isinstance(item, dict) else None
                                if name and isinstance(name, str):
                                    names.append(name.strip())
            except Exception as e:
                logger.error(f"Error reading restaurants.json: {e}")
            return names
        # Load branch names from server API or local file
        site_branches = load_site_branches()
        # Ensure an exact match (case‑sensitive) to enforce the user's instruction.
        if branch_name_input not in site_branches:
            await update.message.reply_text(
                "❌ Bu filial hali qo'shilmagan! Iltimos, filial nomini saytdagi bilan bir xil kiriting."
            )
            # Reset the state to allow the admin to try again.
            context.user_data.pop('adding_branch_admin_name', None)
            return
        # Store the valid branch name and move to the next stage
        context.user_data['branch_name'] = branch_name_input
        context.user_data['adding_branch_admin_name'] = False
        context.user_data['adding_branch_admin_id'] = True
        await update.message.reply_text("Filial uchun admin ID raqamini yuboring:")
        return
    
    elif context.user_data.get('adding_branch_admin_id'):
        # Stage 2: The admin has provided the admin ID for the previously
        # validated branch.  Create or update the KitchenModel accordingly.
        try:
            admin_id = int(text)
        except ValueError:
            await update.message.reply_text("Noto'g'ri ID format! Faqat raqam yuboring.")
            return
        branch_name = context.user_data.get('branch_name')
        if not branch_name:
            # Should not happen; reset state
            context.user_data.pop('adding_branch_admin_id', None)
            await update.message.reply_text("Jarayonni boshidan boshlang.")
            return
        # Check if this branch already has an admin; if so, update it instead of creating a new entry
        existing = next((k for k in kitchens if k.name == branch_name), None)
        if existing:
            existing.admin_id = admin_id
        else:
            kitchens.append(KitchenModel(admin_id=admin_id, name=branch_name))
        # Persist changes
        save_data(KITCHENS_FILE, [k.dict() for k in kitchens])
        await update.message.reply_text(
            f"✅ Filial {branch_name} uchun admin ID {admin_id} belgilandi!"
        )
        # Clear context flags
        context.user_data['adding_branch_admin_id'] = False
        context.user_data.pop('branch_name', None)
        await show_admin_panel(update, context)
        return
    
    elif 'removing_courier' in context.user_data and context.user_data['removing_courier']:
        try:
            courier_id = int(text)
            courier = next((c for c in couriers if c.id == courier_id), None)
            
            if courier:
                couriers.remove(courier)
                save_data(COURIERS_FILE, [c.dict() for c in couriers])
                await update.message.reply_text(f"✅ Kuryer {courier.name} (ID: {courier_id}) o'chirildi!")
            else:
                await update.message.reply_text("❌ Kuryer topilmadi!")
        except ValueError:
            await update.message.reply_text("Noto'g'ri ID format! Faqat raqam yuboring.")
        
        context.user_data['removing_courier'] = False
        await show_admin_panel(update, context)
        return
    
    elif context.user_data.get('removing_branch_admin'):
        # Remove the admin assignment for the specified branch.  We do not
        # delete the branch from the site's data, only from our kitchen list.
        branch_name_input = text.strip()
        # Find the matching KitchenModel entry
        kitchen = next((k for k in kitchens if k.name == branch_name_input), None)
        if kitchen:
            kitchens.remove(kitchen)
            save_data(KITCHENS_FILE, [k.dict() for k in kitchens])
            await update.message.reply_text(
                f"✅ {branch_name_input} filiali uchun admin o'chirildi!"
            )
        else:
            await update.message.reply_text("❌ Filial topilmadi!")
        context.user_data['removing_branch_admin'] = False
        await show_admin_panel(update, context)
        return
    
    elif 'waiting_for_announcement' in context.user_data and context.user_data['waiting_for_announcement']:
        # Send an announcement (any type) to all subscribed users
        for chat_id in list(user_chats):
            try:
                await update.message.copy(chat_id=chat_id)
            except Exception as e:
                logger.error(f"Failed to send announcement to {chat_id}: {e}")
        await update.message.reply_text("✅ Elon barcha foydalanuvchilarga yuborildi!")
        context.user_data['waiting_for_announcement'] = False
        await show_admin_panel(update, context)
        return
    
    elif 'waiting_for_reply' in context.user_data and context.user_data['waiting_for_reply']:
        target_user = context.user_data['replying_to']
        try:
            # Forward admin's response (any message type) to the target user
            await update.message.copy(chat_id=target_user)
            await update.message.reply_text("✅ Javobingiz yuborildi!")
        except Exception as e:
            await update.message.reply_text(f"❌ Javob yuborishda xatolik: {e}")
        
        context.user_data['waiting_for_reply'] = False
        await show_admin_panel(update, context)
        return

    # Handle super admin: adding or removing other admins
    if 'adding_admin' in context.user_data and context.user_data['adding_admin']:
        # Expecting an admin ID
        try:
            admin_id_new = int(text)
            if admin_id_new == SUPER_ADMIN_ID or admin_id_new in EXTRA_ADMIN_IDS:
                await update.message.reply_text("❌ Bu ID allaqachon admin ro'yxatida mavjud.")
            else:
                EXTRA_ADMIN_IDS.add(admin_id_new)
                save_admin_list(list(EXTRA_ADMIN_IDS))
                await update.message.reply_text(f"✅ {admin_id_new} ID raqamli foydalanuvchi admin sifatida qo'shildi!")
        except ValueError:
            await update.message.reply_text("Noto'g'ri ID format! Faqat raqam yuboring.")
        context.user_data['adding_admin'] = False
        await show_admin_panel(update, context)
        return
    if 'removing_admin' in context.user_data and context.user_data['removing_admin']:
        try:
            admin_id_remove = int(text)
            if admin_id_remove == SUPER_ADMIN_ID:
                await update.message.reply_text("❌ Super adminni o'chirish mumkin emas!")
            elif admin_id_remove not in EXTRA_ADMIN_IDS:
                await update.message.reply_text("❌ Bu ID adminlar ro'yxatida topilmadi!")
            else:
                EXTRA_ADMIN_IDS.remove(admin_id_remove)
                save_admin_list(list(EXTRA_ADMIN_IDS))
                await update.message.reply_text(f"✅ {admin_id_remove} ID raqamli admin o'chirildi!")
        except ValueError:
            await update.message.reply_text("Noto'g'ri ID format! Faqat raqam yuboring.")
        context.user_data['removing_admin'] = False
        await show_admin_panel(update, context)
        return

    # If we are not in the middle of a multi‑step admin process, interpret the text as a command from the admin panel.
    # These correspond to the buttons shown in show_admin_panel().
    if text == "➕ Kuryer qo'shish":
        # Start adding a new courier: first ask for the courier's Telegram ID
        context.user_data['adding_courier'] = True
        context.user_data['adding_courier_name'] = False
        await update.message.reply_text("Kuryer ID raqamini yuboring:")
        return
    elif text == "➖ Kuryer o'chirish":
        # Initiate courier removal
        context.user_data['removing_courier'] = True
        await update.message.reply_text("O'chiriladigan kuryer ID raqamini yuboring:")
        return
    elif text == "➕ Filial uchun admin qo'shish":
        # Begin the process of assigning an admin to a branch.  We ask the
        # admin to enter the exact branch name first.  The admin ID will be
        # requested only after validating that the branch exists on the site.
        context.user_data['adding_branch_admin_name'] = True
        context.user_data['adding_branch_admin_id'] = False
        context.user_data.pop('branch_name', None)
        await update.message.reply_text(
            "Filial nomini yuboring\n\n⚠️ Diqqat filial nomini saytdagi bilan bir xil qilib kiriting! Harflarning katta-kichikligi, bo'sh joy va h.k. saytdagi bilan bir xil nom bo'lishi shart! \n\nMasalan saytda Milliy taomlar bo'lsa bu yerga ham aynan Milliy taomlar deb kiriting."
        )
        return
    elif text == "➖ Filial uchun admin o'chirish":
        # Begin the process of removing the admin for a branch.  The admin
        # enters the exact branch name, and if found the corresponding admin
        # assignment is removed.  This does not delete the branch from the
        # website; it only removes the Telegram admin mapping.
        context.user_data['removing_branch_admin'] = True
        await update.message.reply_text(
            "O'chiriladigan filial nomini yuboring (saytdagi bilan bir xil)."
        )
        return
    elif text == "📢 Elon yuborish":
        context.user_data['waiting_for_announcement'] = True
        await update.message.reply_text("📢 Elonni yuboring:")
        return
    elif text == "📊 Statistika":
        await update.message.reply_text(
            f"📊 Statistika:\n👥 Foydalanuvchilar soni: {len(user_chats)}\n🚴 Kuryerlar soni: {len(couriers)}\n🏪 Filiallar soni: {len(kitchens)}"
        )
        return

    # Generate courier info on demand
    if text == "🚚 Kuryer info":
        # Generate statistics for the current day (00:00 to now) in Asia/Tashkent
        try:
            tz = ZoneInfo("Asia/Tashkent")
            now = datetime.now(tz)
            start_dt = datetime.combine(now.date(), time(0, 0), tz)
            end_dt = now
            file_path = await generate_courier_report(start_dt, end_dt)
            # Send the file to the requesting admin only
            try:
                with open(file_path, 'rb') as f:
                    await update.message.reply_document(document=f, filename='kuryerlar.xlsx')
            except Exception as e:
                await update.message.reply_text(f"Faylni yuborishda xatolik: {e}")
            # Remove the file
            try:
                os.remove(file_path)
            except Exception:
                pass
        except Exception as e:
            await update.message.reply_text(f"Ma'lumotlarni yig'ishda xatolik: {e}")
        # After sending, return to admin panel
        await show_admin_panel(update, context)
        return

    # Super admin commands
    if user_id == SUPER_ADMIN_ID:
        if text == "➕ Admin qo'shish":
            context.user_data['adding_admin'] = True
            await update.message.reply_text("Admin ID raqamini yuboring:")
            return
        elif text == "➖ Admin o'chirish":
            context.user_data['removing_admin'] = True
            await update.message.reply_text("O'chiriladigan admin ID raqamini yuboring:")
            return
        elif text == "🧾 Boshqaruv":
            # Generate management report and send as Excel
            try:
                import openpyxl
                from openpyxl import Workbook
            except Exception as e:
                await update.message.reply_text(f"openpyxl kutubxonasi topilmadi: {e}")
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Boshqaruv"
            # Set headers
            ws['A1'] = 'No'
            ws['B1'] = 'Adminlar'
            ws['C1'] = 'Admin ID'
            ws['E1'] = 'Kuryer nomi'
            ws['F1'] = 'Kuryerlar ID'
            ws['H1'] = 'Filial nomlari'
            ws['I1'] = 'Filial admin ID'
            # Populate admins list
            admin_ids_list = [SUPER_ADMIN_ID] + list(EXTRA_ADMIN_IDS)
            # Build list of tuples (name, id) for admins.  Do not show names for any admin IDs defined via
            # the ADMIN_IDS_ENV environment variable (these correspond to super admins defined in code).
            admin_info = []
            for adm_id in admin_ids_list:
                # Determine whether this ID comes from the ADMIN_IDS_ENV list (i.e. built‑in super admins)
                if adm_id in ADMIN_IDS_ENV:
                    # Leave name blank for these entries
                    admin_info.append(("", adm_id))
                else:
                    # Attempt to fetch the chat name; fall back to username or ID if necessary
                    try:
                        chat = await bot.get_chat(adm_id)
                        name = chat.full_name if chat.full_name else (chat.username if chat.username else str(adm_id))
                    except Exception:
                        name = str(adm_id)
                    admin_info.append((name, adm_id))

            # Similarly for couriers
            courier_info = []
            for c in couriers:
                courier_info.append((c.name, c.id))
            # Kitchens
            kitchen_info = []
            for k in kitchens:
                kitchen_info.append((k.name, k.admin_id))
            # Determine maximum number of rows among the three sections
            max_len = max(len(admin_info), len(courier_info), len(kitchen_info))
            for i in range(max_len):
                ws.cell(row=2 + i, column=1, value=i + 1)
                # Admin columns (B and C)
                if i < len(admin_info):
                    ws.cell(row=2 + i, column=2, value=admin_info[i][0])
                    ws.cell(row=2 + i, column=3, value=admin_info[i][1])
                # Courier columns (E and F)
                if i < len(courier_info):
                    ws.cell(row=2 + i, column=5, value=courier_info[i][0])
                    ws.cell(row=2 + i, column=6, value=courier_info[i][1])
                # Kitchen columns (H and I)
                if i < len(kitchen_info):
                    ws.cell(row=2 + i, column=8, value=kitchen_info[i][0])
                    ws.cell(row=2 + i, column=9, value=kitchen_info[i][1])

            # Set column widths according to specification (in pixels).  openpyxl uses character widths, so we
            # approximate conversion: width = (pixels - 5) / 7.  If pixels are too small, we ensure a minimal width.
            def px_to_width(px: int) -> float:
                # guard against very small widths
                width = (px - 5) / 7.0
                return max(width, 1.0)

            column_widths = {
                'A': 30,
                'B': 175,
                'C': 80,
                'D': 15,
                'E': 100,
                'F': 100,
                'G': 15,
                'H': 135,
                'I': 95,
            }
            for col_letter, px in column_widths.items():
                try:
                    ws.column_dimensions[col_letter].width = px_to_width(px)
                except Exception:
                    # If setting width fails for some reason, we continue gracefully
                    pass
            # Save the workbook to a temporary file
            file_path = "/tmp/joriy_royxat.xlsx"
            try:
                wb.save(file_path)
            except Exception as e:
                await update.message.reply_text(f"Faylni saqlashda xatolik: {e}")
                return
            # Send the file
            try:
                with open(file_path, 'rb') as doc:
                    await bot.send_document(chat_id=user_id, document=doc, filename='joriy_royxat.xlsx')
            except Exception as e:
                await update.message.reply_text(f"Faylni yuborishda xatolik: {e}")
            await show_admin_panel(update, context)
            return
    
    # Agar hech qanday holatda bo'lmasa, admin panelini ko'rsatamiz
    await show_admin_panel(update, context)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=PORT)
